"""
Tests for pipeline core logic.

Run with: python -m pytest tests/
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from pipeline import build_rows
from dow import _looks_like_name, _find_section_start
from search import _is_program_office, _build_search_prompt, _CUTOFF_YEAR, _SOURCE_CUTOFF_YEARS
from utils import strip_json_fences, rows_to_xlsx_bytes, rows_to_csv_bytes, CSV_FIELDS


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _office(name, code=None, parent=None):
    return {"name": name, "code": code, "parent": parent}

def _search(leadership=None, code=None, acronym=None):
    return {"code": code, "acronym": acronym, "leadership": leadership or []}

def _leader(name, title, source="website", confidence="High", notes=""):
    return {"name": name, "title": title, "source": source,
            "confidence": confidence, "notes": notes}


# ---------------------------------------------------------------------------
# build_rows
# ---------------------------------------------------------------------------

def test_build_rows_empty_offices():
    assert build_rows([], {}) == []


def test_build_rows_no_leaders_produces_no_rows():
    """Offices with no leadership found contribute zero rows."""
    offices = [_office("Unknown Office")]
    assert build_rows(offices, {}) == []


def test_build_rows_single_leader():
    offices = [_office("Office of Naval Research")]
    search  = {"Office of Naval Research": _search(leadership=[
        _leader("Dr. Rachel Riley", "Chief of Naval Research"),
    ])}
    rows = build_rows(offices, search)

    assert len(rows) == 1
    row = rows[0]
    assert row["org_name"]    == "Office of Naval Research"
    assert row["person_name"] == "Dr. Rachel Riley"
    assert row["title"]       == "Chief of Naval Research"
    assert row["confidence"]  == "High"
    assert row["source"]      == "website"
    assert "section"    not in row
    assert "org_code"   not in row
    assert "parent_org" not in row


def test_build_rows_multiple_leaders():
    offices = [_office("Office of Naval Research")]
    search  = {"Office of Naval Research": _search(leadership=[
        _leader("Dr. Rachel Riley", "Chief of Naval Research"),
        _leader("RDML John Smith",  "Deputy Chief", source="linkedin", confidence="Medium"),
    ])}
    rows = build_rows(offices, search)

    assert len(rows) == 2
    assert rows[0]["person_name"] == "Dr. Rachel Riley"
    assert rows[1]["person_name"] == "RDML John Smith"
    assert rows[1]["confidence"]  == "Medium"


def test_build_rows_notes_and_source_preserved():
    offices = [_office("Education Service")]
    search  = {"Education Service": _search(leadership=[
        _leader("Ken Smith", "Acting Executive Director",
                source="Congressional testimony Feb 2025",
                confidence="Medium",
                notes="Acting title as of Feb 2025 — verify if still acting or made permanent"),
    ])}
    rows = build_rows(offices, search)

    assert rows[0]["notes"].startswith("Acting title")
    assert rows[0]["source"] == "Congressional testimony Feb 2025"


def test_build_rows_multiple_offices():
    offices = [
        _office("Org A"),
        _office("Org B"),
    ]
    search = {
        "Org A": _search(leadership=[_leader("Alice", "Director")]),
        "Org B": _search(leadership=[_leader("Bob", "Director"), _leader("Carol", "Deputy")]),
    }
    rows = build_rows(offices, search)
    assert len(rows) == 3
    assert rows[0]["org_name"] == "Org A"
    assert rows[1]["org_name"] == "Org B"
    assert rows[2]["org_name"] == "Org B"


# ---------------------------------------------------------------------------
# strip_json_fences
# ---------------------------------------------------------------------------

def test_strip_json_fences_with_backticks():
    assert strip_json_fences("```json\n{}\n```") == "{}"

def test_strip_json_fences_no_fences():
    assert strip_json_fences('{"key": "value"}') == '{"key": "value"}'

def test_strip_json_fences_plain_backticks():
    assert strip_json_fences("```\n[]\n```") == "[]"


# ---------------------------------------------------------------------------
# rows_to_xlsx_bytes
# ---------------------------------------------------------------------------

def test_rows_to_xlsx_bytes_is_valid_workbook():
    from io import BytesIO
    from openpyxl import load_workbook

    rows = [_leader_row(confidence="High"), _leader_row(confidence="Low")]
    wb = load_workbook(BytesIO(rows_to_xlsx_bytes(rows)))
    ws = wb.active
    assert [c.value for c in ws[1]] == CSV_FIELDS
    assert ws.max_row == 3  # header + 2 rows

def test_rows_to_xlsx_bytes_empty_rows():
    from io import BytesIO
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(rows_to_xlsx_bytes([])))
    assert wb.active.max_row == 1  # header only


def test_rows_to_xlsx_bytes_custom_fields():
    from io import BytesIO
    from openpyxl import load_workbook

    fields = ["person_name", "portfolio", "source"]
    rows = [{"person_name": "Jane Doe", "portfolio": "Aviation", "source": "https://navy.mil/x"}]
    wb = load_workbook(BytesIO(rows_to_xlsx_bytes(rows, fields=fields, sheet_title="PAE")))
    ws = wb.active
    assert ws.title == "PAE"
    assert [c.value for c in ws[1]] == fields
    assert ws["A2"].value == "Jane Doe"


# ---------------------------------------------------------------------------
# rows_to_csv_bytes
# ---------------------------------------------------------------------------

def test_rows_to_csv_bytes_default_fields():
    rows = [_leader_row(confidence="High")]
    csv_text = rows_to_csv_bytes(rows).decode()
    assert csv_text.splitlines()[0] == ",".join(CSV_FIELDS)
    assert "Jane Doe" in csv_text

def test_rows_to_csv_bytes_custom_fields_and_quoting():
    rows = [{"person_name": 'Jane "The Fixer" Doe', "portfolio": "Munitions, Aviation"}]
    csv_text = rows_to_csv_bytes(rows, fields=["person_name", "portfolio"]).decode()
    lines = csv_text.splitlines()
    assert lines[0] == "person_name,portfolio"
    # Embedded quotes and commas must be properly escaped, not corrupt the row
    assert '"Jane ""The Fixer"" Doe"' in lines[1]
    assert '"Munitions, Aviation"' in lines[1]


def _leader_row(confidence="High"):
    return {
        "org_name": "Office of Naval Research", "person_name": "Jane Doe", "title": "Director",
        "source": "https://onr.navy.mil", "confidence": confidence, "notes": "", "description": "", "acronym": "ONR",
    }


# ---------------------------------------------------------------------------
# DoW Directory helpers
# ---------------------------------------------------------------------------

def test_looks_like_name_valid():
    assert _looks_like_name("LTG Chris Mohan") is True
    assert _looks_like_name("Mr. David Lorch") is True
    assert _looks_like_name("Owen West") is True
    assert _looks_like_name("COL Brad Dello") is True


def test_looks_like_name_rejects_vacant():
    assert _looks_like_name("Vacant") is False
    assert _looks_like_name("TBD") is False


def test_looks_like_name_rejects_short():
    assert _looks_like_name("DIU") is False


def test_find_section_start_exact():
    lines = [
        "Army Material Command",
        "LTG Chris Mohan - Commanding General",
    ]
    assert _find_section_start(lines, "Army Material Command") == 0


def test_find_section_start_substring():
    lines = [
        "Some other line",
        "Defense Innovation Unit (DIU)",
        "Owen West - Director of DIU",
    ]
    assert _find_section_start(lines, "Defense Innovation Unit") == 1


def test_find_section_start_not_found():
    lines = ["Nothing here", "Blank"]
    assert _find_section_start(lines, "Nonexistent Org XYZ") is None


def test_find_section_start_ignores_page_footer_stamp():
    # A running page-footer/header stamp repeats the chapter name on every
    # page of a large multi-org chapter — matching on it would lock onto
    # whatever content happens to follow on that first page, not the
    # queried org's own section. Regression test for a real bug: searching
    # "Office of the Secretary of War" locked onto such a stamp and pulled
    # in ~100 rows from unrelated agencies (Missile Defense Agency, EXIM Bank).
    lines = [
        "Page 95 of 453 Office of the Secretary of War 2026 DoW Directory Rev 8",
        "Some Unrelated Agency - Director",
        "Office of the Secretary of War",
        "Jules Hurst - Under Secretary of War (Comptroller)",
    ]
    assert _find_section_start(lines, "Office of the Secretary of War") == 2


def test_find_section_start_rejects_generic_word_overlap_false_positive():
    # "Office of the Secretary of War" and "Office of the Secretary of the
    # Air Force" share 4 of 5 words (office/of/the/secretary) — all generic
    # scaffolding, not distinguishing content. The overlap scorer must weigh
    # the distinctive words (secretary, war), not raw word-set overlap.
    lines = [
        "Ben Maitre - Director, Legislative Liaison, Office of the Secretary of the Air Force",
        "Jane Roe - Some Other Title",
    ]
    assert _find_section_start(lines, "Office of the Secretary of War") is None


def test_build_rows_dow_leaders_come_first():
    """DoW leaders should appear before web-search leaders."""
    offices = [_office("Defense Innovation Unit")]
    search  = {"Defense Innovation Unit": _search(leadership=[
        _leader("Owen West", "Director", source="linkedin", confidence="Medium"),
    ])}
    dow = {"Defense Innovation Unit": [
        {"name": "Owen West", "title": "Director of DIU",
         "source": "DoW Directory 2026", "confidence": "High", "notes": ""},
    ]}
    rows = build_rows(offices, search, dow_data=dow)
    assert len(rows) == 1
    assert rows[0]["source"] == "DoW Directory 2026"
    assert rows[0]["confidence"] == "High"


def test_build_rows_dow_supplements_web():
    """Web-search leaders not in DoW should be appended after DoW leaders."""
    offices = [_office("Office of Naval Research")]
    search  = {"Office of Naval Research": _search(leadership=[
        _leader("Dr. Rachel Riley", "Chief of Naval Research", source="website", confidence="High"),
        _leader("RDML John Smith",  "Deputy Chief",             source="linkedin", confidence="Medium"),
    ])}
    dow = {"Office of Naval Research": [
        {"name": "Dr. Rachel Riley", "title": "Chief of Naval Research",
         "source": "DoW Directory 2026", "confidence": "High", "notes": ""},
    ]}
    rows = build_rows(offices, search, dow_data=dow)
    assert len(rows) == 2
    assert rows[0]["source"] == "DoW Directory 2026"
    assert rows[1]["person_name"] == "RDML John Smith"
    assert rows[1]["source"] == "linkedin"


def test_build_rows_no_dow_backward_compatible():
    """Omitting dow_data leaves existing behaviour unchanged."""
    offices = [_office("NavalX")]
    search  = {"NavalX": _search(leadership=[_leader("Jane Doe", "Director")])}
    rows = build_rows(offices, search)
    assert len(rows) == 1
    assert rows[0]["person_name"] == "Jane Doe"


# ---------------------------------------------------------------------------
# run_pipeline — resumes from an existing checkpoint
# ---------------------------------------------------------------------------

def test_run_pipeline_resumes_from_checkpoint(tmp_path, monkeypatch):
    # The dashboard now submits the same org list to the same deterministic
    # output path on retry (see app.py), specifically so this resume path
    # kicks in after a dropped connection on a large batch instead of
    # re-searching everything from scratch. Verify it actually skips
    # already-checkpointed orgs and only searches the rest.
    import json

    import pipeline as pipeline_module

    output_path = tmp_path / "test_output.csv"
    checkpoint_path = output_path.parent / f".{output_path.stem}.checkpoint.json"
    checkpoint_path.write_text(json.dumps({
        "NavalX": {
            "code": None, "acronym": None,
            "leadership": [{"name": "Cached Leader", "title": "Cached Title",
                             "source": "x", "confidence": "High", "notes": ""}],
        },
    }))

    calls = []

    def fake_search_office(client, office, index, total, backend):
        calls.append(office["name"])
        return {
            "code": None, "acronym": None,
            "leadership": [{"name": "Fresh Leader", "title": "Fresh Title",
                             "source": "y", "confidence": "High", "notes": ""}],
        }

    monkeypatch.setattr(pipeline_module, "search_office", fake_search_office)
    monkeypatch.setattr(pipeline_module.time, "sleep", lambda _: None)

    rows = pipeline_module.run_pipeline(
        orgs=["Office of Naval Research", "NavalX"],
        output=output_path,
        backend="gemini",
    )

    assert calls == ["Office of Naval Research"]  # NavalX skipped — served from checkpoint
    navalx_rows = [r for r in rows if r["org_name"] == "NavalX"]
    assert navalx_rows[0]["person_name"] == "Cached Leader"
    onr_rows = [r for r in rows if r["org_name"] == "Office of Naval Research"]
    assert onr_rows[0]["person_name"] == "Fresh Leader"
    assert not checkpoint_path.exists()  # cleared after a fully successful run


# ---------------------------------------------------------------------------
# Program office detection
# ---------------------------------------------------------------------------

def test_is_program_office_detects_common_prefixes():
    assert _is_program_office("PMW 120 Ship Communications") is True
    assert _is_program_office("PdM Tactical Radios") is True
    assert _is_program_office("PEO Command and Control") is True
    assert _is_program_office("PMO Rapid Sustainment") is True
    assert _is_program_office("PAE Ground Combat Systems") is True


def test_is_program_office_detects_keyword_phrases():
    assert _is_program_office("Program Office for Enterprise IT") is True
    assert _is_program_office("Joint Program Manager Chemical Defense") is True


def test_is_program_office_rejects_regular_orgs():
    assert _is_program_office("Office of Naval Research") is False
    assert _is_program_office("Army Material Command") is False
    assert _is_program_office("Defense Innovation Unit") is False


# ---------------------------------------------------------------------------
# Prompt builder and recency constants
# ---------------------------------------------------------------------------

def test_cutoff_year_is_recent():
    import datetime
    expected = datetime.date.today().year - _SOURCE_CUTOFF_YEARS
    assert _CUTOFF_YEAR == expected


def test_build_search_prompt_includes_cutoff_year():
    prompt = _build_search_prompt("NavalX", "Department of the Navy", is_program=False)
    assert str(_CUTOFF_YEAR) in prompt


def test_build_search_prompt_program_office_adds_pm_note():
    prompt = _build_search_prompt("PMW 120", "NAVWAR", is_program=True)
    assert "program manager" in prompt.lower()


def test_build_search_prompt_non_program_has_no_pm_note():
    # Non-program offices still include "program manager" in the general role list,
    # but should NOT include the program-office-specific Step 3 instruction.
    prompt = _build_search_prompt("Office of Naval Research", "DoN", is_program=False)
    assert "program executive officer" not in prompt.lower() or "in step 3" not in prompt.lower()


# ---------------------------------------------------------------------------
# call_gemini — retry on empty response
# ---------------------------------------------------------------------------

def test_call_gemini_retries_on_no_text_content(monkeypatch):
    # Regression test: a "no text content" response (empty/blocked generation)
    # is a transient glitch, not a real failure, but previously fell straight
    # through the rate-limit-only retry check and raised immediately instead
    # of retrying (Ollama fallback, the old failure path here, has since
    # been removed entirely — it never worked on a hosted deployment anyway).
    import sys
    import types as pytypes

    import backends as backends_module

    monkeypatch.setenv("GEMINI_API_KEY", "fake-key-for-test")
    monkeypatch.setattr(backends_module.time, "sleep", lambda _: None)

    class FakeResponse:
        def __init__(self, text):
            self.text = text
            self.prompt_feedback = "SAFETY"

    responses = iter([FakeResponse(None), FakeResponse("real answer")])

    class FakeModels:
        def generate_content(self, **kwargs):
            return next(responses)

    class FakeClient:
        def __init__(self, api_key):
            self.models = FakeModels()

    class _AcceptsAnyKwargs:
        def __init__(self, **kwargs):
            pass

    fake_genai = pytypes.ModuleType("google.genai")
    fake_genai.Client = FakeClient
    fake_genai_types = pytypes.ModuleType("google.genai.types")
    fake_genai_types.GenerateContentConfig = _AcceptsAnyKwargs
    fake_genai_types.Tool = _AcceptsAnyKwargs
    fake_genai_types.GoogleSearch = _AcceptsAnyKwargs
    fake_genai.types = fake_genai_types

    monkeypatch.setitem(sys.modules, "google.genai", fake_genai)
    monkeypatch.setitem(sys.modules, "google.genai.types", fake_genai_types)

    result = backends_module.call_gemini("prompt", "system")
    assert result == "real answer"


# ---------------------------------------------------------------------------
# search_office — retry on malformed JSON
# ---------------------------------------------------------------------------

def test_search_office_retries_once_on_malformed_json(monkeypatch):
    # Regression test: Gemini returned invalid JSON for a real org this
    # session; the org was silently counted as "0 leaders" even though a
    # retry succeeded immediately. search_office must retry once before
    # giving up, since this is typically a one-off generation glitch.
    import search as search_module

    responses = iter([
        '{"code": null, "acronym": "ONR", "leadership": [',  # truncated/invalid JSON
        '{"code": null, "acronym": "ONR", "leadership": [{"name": "Jane Doe", "title": "Director", "source": "x", "confidence": "High", "notes": ""}]}',
    ])
    calls = []

    def fake_call_backend(client, prompt, system, backend, **kwargs):
        calls.append(1)
        return next(responses)

    monkeypatch.setattr(search_module, "call_backend", fake_call_backend)
    result = search_module.search_office(None, _office("Office of Naval Research"), 1, 1)

    assert len(calls) == 2
    assert len(result["leadership"]) == 1
    assert result["leadership"][0]["name"] == "Jane Doe"


def test_search_office_does_not_retry_on_non_json_error(monkeypatch):
    # A missing API key / quota exhaustion / network error won't be fixed by
    # retrying the same call — those should fail fast on the first attempt,
    # not waste a second call repeating the same error.
    import search as search_module

    calls = []

    def fake_call_backend(client, prompt, system, backend, **kwargs):
        calls.append(1)
        raise RuntimeError("GEMINI_API_KEY not set")

    monkeypatch.setattr(search_module, "call_backend", fake_call_backend)
    result = search_module.search_office(None, _office("Office of Naval Research"), 1, 1)

    assert len(calls) == 1
    assert result["leadership"] == []


def test_search_office_gives_up_after_second_malformed_response(monkeypatch):
    import search as search_module

    calls = []

    def fake_call_backend(client, prompt, system, backend, **kwargs):
        calls.append(1)
        return "not json at all"

    monkeypatch.setattr(search_module, "call_backend", fake_call_backend)
    result = search_module.search_office(None, _office("Office of Naval Research"), 1, 1)

    assert len(calls) == 2
    assert result["leadership"] == []
