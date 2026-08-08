"""
utils.py — Shared infrastructure for the org-pipeline tools.

Provides:
  - TypedDicts that define the data contracts between pipeline stages
  - Environment loading and Anthropic client setup
  - Claude API wrappers with automatic retry and exponential backoff
  - Shared CSV schema and writer
  - JSON fence stripping
"""
from __future__ import annotations

__all__ = [
    # TypedDicts
    "OrgEntry",
    "SearchResult",
    "PdfProgram",
    "PdfOfficeData",
    # Constants
    "CSV_FIELDS",
    "INTER_REQUEST_DELAY",
    # Environment
    "load_env",
    "get_anthropic_client",
    "get_model",
    # Encoding
    "encode_image",
    # Claude
    "claude",
    "claude_with_search",
    # CSV
    "write_csv",
    "rows_to_xlsx_bytes",
    # JSON
    "strip_json_fences",
]

import base64
import csv
import logging
import os
import re
import time
from pathlib import Path
from typing import TypedDict

import anthropic

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# TypedDicts — explicit data contracts between pipeline stages
# ---------------------------------------------------------------------------

class OrgEntry(TypedDict):
    """A single organizational unit extracted from an org chart image."""

    name: str
    code: str | None
    parent: str | None


class SearchResult(TypedDict):
    """Data returned by a web search for one office.

    Only fields actually returned by the search prompt are included here.
    website/phone/email were previously tracked but never returned by the
    prompt or written to CSV — removed to eliminate silent data drift.
    """

    code: str | None
    acronym: str | None
    leadership: list[dict] | None


class PdfProgram(TypedDict):
    """A named program found within a Functional Organization Manual."""

    name: str
    acronym: str | None
    description: str | None


class PdfOfficeData(TypedDict):
    """Description and sub-programs extracted from a FOM for one office."""

    description: str | None
    programs: list[PdfProgram]


# ---------------------------------------------------------------------------
# Environment
# ---------------------------------------------------------------------------

def load_env(env_path: str | Path | None = None) -> None:
    """Load ``key=value`` pairs from a .env file into ``os.environ``.

    Already-set variables are never overwritten, so shell exports always
    take precedence over the .env file.

    Args:
        env_path: Path to the .env file. Defaults to ``.env`` in the same
                  directory as this module.
    """
    path = Path(env_path) if env_path else Path(__file__).parent / ".env"
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if value and "paste_your_key_here" not in value:
            os.environ.setdefault(key, value)


def get_anthropic_client() -> anthropic.Anthropic:
    """Load .env and return an authenticated Anthropic client.

    The active model can be overridden at runtime without modifying source::

        export ANTHROPIC_MODEL=claude-opus-4-8

    Raises:
        SystemExit: If ``ANTHROPIC_API_KEY`` is absent from both the
                    environment and the .env file.
    """
    load_env()
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise SystemExit(
            "Error: ANTHROPIC_API_KEY is not set.\n"
            "  Option 1: export ANTHROPIC_API_KEY=sk-ant-...\n"
            "  Option 2: add it to .env  (see .env.example)"
        )
    return anthropic.Anthropic(api_key=api_key)


# ---------------------------------------------------------------------------
# Model selection
# ---------------------------------------------------------------------------

_DEFAULT_MODEL = "claude-sonnet-4-6"


def get_model() -> str:
    """Active model ID; override with ANTHROPIC_MODEL env var."""
    return os.environ.get("ANTHROPIC_MODEL", _DEFAULT_MODEL)


# ---------------------------------------------------------------------------
# Image encoding
# ---------------------------------------------------------------------------

def encode_image(image_path: str | Path) -> tuple[str, str]:
    """Base64-encode a local image for use in a Claude vision content block.

    Args:
        image_path: Path to a PNG, JPG, GIF, or WebP file.

    Returns:
        ``(base64_data, media_type)`` ready for the Claude API image source dict.
    """
    path = Path(image_path)
    media_type_map = {
        ".jpg":  "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png":  "image/png",
        ".gif":  "image/gif",
        ".webp": "image/webp",
    }
    media_type = media_type_map.get(path.suffix.lower(), "image/png")
    return base64.standard_b64encode(path.read_bytes()).decode(), media_type


# ---------------------------------------------------------------------------
# API layer — retry with exponential backoff
# ---------------------------------------------------------------------------

_MAX_RETRIES: int = 3

# Pause between consecutive office searches. Keeps aggregate request rate
# well below Anthropic's per-minute token limits on busy orgs.
INTER_REQUEST_DELAY: float = 1.0


def _messages_create(client: anthropic.Anthropic, **kwargs) -> anthropic.types.Message:
    """Call ``client.messages.create`` with exponential backoff on transient errors.

    Retries up to ``_MAX_RETRIES`` times on:

    - ``RateLimitError`` (HTTP 429) — back off and retry.
    - ``APIStatusError`` with a 5xx status — server fault, safe to retry.

    Client errors (4xx, excluding 429) are raised immediately because
    retrying them will not produce a different result.

    Args:
        client: Authenticated Anthropic client.
        **kwargs: Forwarded verbatim to ``messages.create``.

    Returns:
        The API response ``Message`` object.

    Raises:
        anthropic.APIError: After all retries are exhausted, or immediately
                            for non-retryable errors.
    """
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            return client.messages.create(**kwargs)
        except anthropic.RateLimitError:
            if attempt == _MAX_RETRIES:
                raise
            delay = 2.0 ** attempt
            logger.warning(
                "Rate limit hit — retrying in %.0fs (attempt %d/%d)",
                delay, attempt, _MAX_RETRIES,
            )
            time.sleep(delay)
        except anthropic.APIStatusError as exc:
            if attempt == _MAX_RETRIES or exc.status_code < 500:
                raise
            delay = 2.0 ** attempt
            logger.warning(
                "API %d error — retrying in %.0fs (attempt %d/%d)",
                exc.status_code, delay, attempt, _MAX_RETRIES,
            )
            time.sleep(delay)

    raise RuntimeError("unreachable")  # keeps type checkers happy


# ---------------------------------------------------------------------------
# Claude helpers
# ---------------------------------------------------------------------------

def claude(
    client: anthropic.Anthropic,
    messages: list[dict],
    system: str = "",
    model: str | None = None,
    max_tokens: int = 4096,
) -> str:
    """Make a single-turn Claude API call and return the first text block.

    Args:
        client:     Authenticated Anthropic client.
        messages:   Conversation in the Claude API message format.
        system:     Optional system prompt.
        model:      Model ID override; defaults to ``get_model()``.
        max_tokens: Maximum tokens in the response.

    Returns:
        Text content of the first response block.
    """
    kwargs: dict = {
        "model":      model or get_model(),
        "max_tokens": max_tokens,
        "messages":   messages,
    }
    if system:
        kwargs["system"] = system
    response = _messages_create(client, **kwargs)
    if not response.content:
        raise RuntimeError(f"Empty response content (stop_reason={response.stop_reason!r})")
    return response.content[0].text


def claude_with_search(
    client: anthropic.Anthropic,
    prompt: str,
    system: str = "",
    max_search_uses: int = 5,
) -> str:
    """Call Claude with the built-in web search tool and return the final answer.

    Implements the tool-use agentic loop: Claude may call ``web_search``
    multiple times before producing a final ``end_turn`` response. Each
    intermediate API call goes through ``_messages_create``, so it inherits
    the same retry / backoff behaviour as a plain ``claude()`` call.

    Args:
        client:          Authenticated Anthropic client.
        prompt:          User message.
        system:          Optional system prompt.
        max_search_uses: Maximum web searches Claude may perform per call.

    Returns:
        Concatenated text from all response blocks in the final turn.
    """
    tools = [{"type": "web_search_20250305", "name": "web_search", "max_uses": max_search_uses}]
    messages: list[dict] = [{"role": "user", "content": prompt}]
    kwargs: dict = {
        "model":      get_model(),
        "max_tokens": 2048,
        "tools":      tools,
        "messages":   messages,
    }
    if system:
        kwargs["system"] = system

    while True:
        response = _messages_create(client, **kwargs)
        text_blocks = [b.text for b in response.content if hasattr(b, "text")]

        if response.stop_reason == "end_turn":
            return " ".join(text_blocks)

        tool_uses = [b for b in response.content if b.type == "tool_use"]
        if not tool_uses:
            # Model stopped without a tool call and without end_turn —
            # return whatever text was produced rather than looping forever.
            return " ".join(text_blocks)

        # Append the assistant turn and acknowledge each tool invocation.
        # web_search results are injected server-side; the tool_result entries
        # here satisfy the API's conversational turn-structure requirements.
        messages.append({"role": "assistant", "content": response.content})
        messages.append({
            "role": "user",
            "content": [
                {
                    "type":        "tool_result",
                    "tool_use_id": tu.id,
                    "content":     getattr(tu, "content", "") or "",
                }
                for tu in tool_uses
            ],
        })


# ---------------------------------------------------------------------------
# CSV output
# ---------------------------------------------------------------------------

# Canonical column order for CRM import. Defined once so that all scripts
# produce files with an identical schema that can be concatenated directly.
CSV_FIELDS: list[str] = [
    "org_name", "person_name", "title", "source", "confidence", "notes", "description", "acronym",
]


def write_csv(rows: list[dict], output_path: str | Path) -> None:
    """Write rows to the canonical CRM CSV. Extra dict keys are silently ignored."""
    path = Path(output_path)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    logger.info("Wrote %d rows → %s", len(rows), path)


# Soft fill colors per confidence level — legible on-screen scan aid for
# non-technical CRM users, not just a raw CSV of undifferentiated rows.
_CONFIDENCE_FILLS = {
    "High":   "DCEEE0",
    "Medium": "FBEFD7",
    "Low":    "F1E2E0",
}
_COLUMN_WIDTHS = {
    "org_name": 32, "person_name": 22, "title": 38, "source": 42,
    "confidence": 12, "notes": 46, "description": 46, "acronym": 12,
}


def rows_to_xlsx_bytes(rows: list[dict]) -> bytes:
    """Build a formatted .xlsx workbook from CSV rows and return its bytes.

    Bold frozen header, sized columns, an autofilter, and a soft color fill
    on the confidence column — meant to open directly in Excel or Google
    Sheets and be immediately scannable without any CSV import step.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leadership"

    header_fill = PatternFill("solid", fgColor="211F1C")
    header_font = Font(color="F9F9F8", bold=True)
    ws.append(CSV_FIELDS)
    for col_idx, _ in enumerate(CSV_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    confidence_col = CSV_FIELDS.index("confidence") + 1
    for row in rows:
        ws.append([row.get(field, "") for field in CSV_FIELDS])
        r = ws.max_row
        fill_color = _CONFIDENCE_FILLS.get(row.get("confidence"))
        if fill_color:
            ws.cell(row=r, column=confidence_col).fill = PatternFill("solid", fgColor=fill_color)
        for col_idx in range(1, len(CSV_FIELDS) + 1):
            ws.cell(row=r, column=col_idx).alignment = Alignment(vertical="top", wrap_text=False)

    for col_idx, field in enumerate(CSV_FIELDS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _COLUMN_WIDTHS.get(field, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CSV_FIELDS))}{ws.max_row}"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------

def strip_json_fences(text: str) -> str:
    """Remove ```json fences that Claude adds despite being told not to."""
    return re.sub(r"```(?:json)?", "", text).strip().rstrip("`").strip()
